from datetime import timedelta, date
import openpyxl
import os

wb = openpyxl.load_workbook(".\Thesis\TrainingData.xlsx")
ws = wb.active

newWb = openpyxl.Workbook()
newWs = newWb.create_sheet("TrainingDataset(2021)")

lastRow = ws.max_row

finalizedData = [[]]
for row in range(lastRow):
    # Prevent create empty lists
    if len(finalizedData[-1]) != 0:
        finalizedData.append([])

    # Loop through each row
    alphaCol = ['B','C','D','E','F','G','H','I','J','K','L']
    alphaCol.reverse()
    for letter in alphaCol:
        # Skip headers
        if row==0:
            finalizedData[-1].insert(0, ws[f'{letter}{row+1}'].value)
            continue
        # Start iterating data
        if ws[f'{letter}{row+1}'].value == None:
            continue

        if letter in ['B']:
            monthDay = ws[f'{letter}{row+1}'].value.split('-')
            newDate = date(year=2021, month=int(monthDay[0]), day=int(monthDay[1]))
            finalizedData[-1].insert(0, newDate)
        elif letter in ['D', 'E', 'F', 'H', 'I', 'J', 'K']:
            completeNum = ""
            for char in ws[f'{letter}{row+1}'].value:
                if char.isnumeric() or char=='.':
                    completeNum += char
            finalizedData[-1].insert(0, float(completeNum))
        else:
            finalizedData[-1].insert(0, ws[f'{letter}{row+1}'].value)

for idxRow, row in enumerate(finalizedData):
    for idxCol, col in enumerate(row):
        newWs.cell(column=idxCol+1, row=idxRow+1).value = col

newWb.save('C:\Users\Asus\Documents\CleanedData.xlsx')
